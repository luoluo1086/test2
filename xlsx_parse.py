import os
from openpyxl import load_workbook


class XlsxParse:
    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(path)
        self.sheet = self.wb.active

    def get_doc_info(self):
        doc_info = {}
        keys = []
        values = []
        for row in self.sheet.iter_rows(min_row=2, max_row=2, values_only=True):
            for i, key in enumerate(row):
                keys.append(key)
        for row in self.sheet.iter_rows(min_row=3, max_row=3, values_only=True):
            for i, value in enumerate(row):
                values.append(value)
        for i in range(len(keys)):
            doc_info[keys[i]] = values[i]
        return doc_info
    
    def get_requirement_info(self):
        requirement_info = []
        keys = []
        for row in self.sheet.iter_rows(min_row=5, max_row=5, values_only=True):
            for i, key in enumerate(row):
                keys.append(key)
        for row in self.sheet.iter_rows(min_row=6, values_only=True):
            values = []
            for i, value in enumerate(row):
                values.append(value)
            requirement = dict(zip(keys, values))
            requirement_info.append(requirement)
        return requirement_info


if __name__ == '__main__':
    xlsx_parse = XlsxParse('./template.xlsx')
    doc_info = xlsx_parse.get_doc_info()
    print(doc_info)
    requirement_info = xlsx_parse.get_requirement_info()
    print(requirement_info)